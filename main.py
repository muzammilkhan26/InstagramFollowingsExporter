import os
import io
import json
import time
import math
import queue
import shutil
import string
import random
import threading
import datetime as dt
from dataclasses import dataclass, asdict

import requests
from requests.cookies import RequestsCookieJar

# GUI
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Data / Export
import pandas as pd

# Excel exports
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except Exception:
    Workbook = None

HEADERS = {
    # Mobile UA generally works reliably with i.instagram.com endpoints
    "User-Agent": "Instagram 219.0.0.12.117 Android",
    "Accept": "*/*",
    "Accept-Language": "en-US,en;q=0.9",
    "X-IG-App-ID": "936619743392459"  # public web app id; helps some web endpoints
}

INSTAGRAM_BASE = "https://i.instagram.com"


# ----------------------------- Utilities -----------------------------

def read_netscape_cookies_txt(path: str) -> RequestsCookieJar:
    jar = RequestsCookieJar()
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            # Netscape format: domain, flag, path, secure, expiration, name, value
            parts = line.split('\t')
            if len(parts) < 7:
                # some exporters use spaces
                parts = line.split()
            if len(parts) < 7:
                continue
            domain, _flag, path, secure, _exp, name, value = parts[-7:]
            secure = secure.lower() == 'true'
            jar.set(name, value, domain=domain, path=path, secure=secure)
    return jar


def safe_filename(stem: str) -> str:
    keep = f"-_.() {string.ascii_letters}{string.digits}"
    return ''.join(c for c in stem if c in keep).strip().replace(' ', '_')


@dataclass
class Followee:
    username: str
    profile_link: str
    full_name: str | None = None
    is_verified: bool | None = None
    is_private: bool | None = None
    user_id: str | None = None
    profile_pic_url: str | None = None
    followers: int | None = None
    followings: int | None = None
    bio: str | None = None


# ----------------------------- Instagram Client -----------------------------

class IGClient:
    def __init__(self, cookies_path: str):
        self.s = requests.Session()
        self.s.headers.update(HEADERS)
        self.s.cookies = read_netscape_cookies_txt(cookies_path)

    def _get(self, url, **kwargs):
        r = self.s.get(url, timeout=30, **kwargs)
        return r

    def validate_session(self) -> tuple[bool, str | None]:
        # Light-weight call that requires auth and returns current user info
        url = f"{INSTAGRAM_BASE}/api/v1/accounts/current_user/"
        r = self._get(url)
        if r.status_code == 200:
            try:
                data = r.json()
                username = data.get('user', {}).get('username')
                return True, username
            except Exception:
                return True, None
        return False, None

    def get_user_id(self, username: str) -> str:
        url = f"{INSTAGRAM_BASE}/api/v1/users/web_profile_info/?username={username}"
        r = self._get(url)
        if r.status_code != 200:
            raise RuntimeError("Cookies invalid ya expired lag rahi hain (profile_info fail).")
        data = r.json()
        try:
            return data["data"]["user"]["id"]
        except Exception:
            raise RuntimeError("Username se user_id parse nahi ho paya. Shayad account private/changed.")

    def get_followings(self, user_id: str, fetch_details=False, throttle_sec=0.4):
        followees: list[Followee] = []
        next_max_id = None
        while True:
            url = f"{INSTAGRAM_BASE}/api/v1/friendships/{user_id}/following/?count=200"
            if next_max_id:
                url += f"&max_id={next_max_id}"
            r = self._get(url)
            if r.status_code != 200:
                raise RuntimeError(f"Followings fetch fail: HTTP {r.status_code}")
            js = r.json()
            users = js.get('users', [])
            for u in users:
                fe = Followee(
                    username=u.get('username'),
                    profile_link=f"https://instagram.com/{u.get('username')}",
                    full_name=u.get('full_name'),
                    is_verified=u.get('is_verified'),
                    is_private=u.get('is_private'),
                    user_id=str(u.get('pk')) if u.get('pk') is not None else None,
                    profile_pic_url=u.get('profile_pic_url') or u.get('profile_pic_url_hd')
                )
                followees.append(fe)
            next_max_id = js.get('next_max_id')
            if not next_max_id:
                break
            time.sleep(0.25)

        if fetch_details:
            # Enrich with counts + bio; one call per user (slow!)
            for fe in followees:
                try:
                    if not fe.user_id:
                        continue
                    info_url = f"{INSTAGRAM_BASE}/api/v1/users/{fe.user_id}/info/"
                    r = self._get(info_url)
                    if r.status_code != 200:
                        continue
                    info = r.json().get('user', {})
                    fe.followers = info.get('follower_count')
                    fe.followings = info.get('following_count')
                    fe.bio = info.get('biography')
                    # Prefer HD pic if available
                    fe.profile_pic_url = info.get('hd_profile_pic_url_info', {}).get('url') or fe.profile_pic_url
                except Exception:
                    pass
                time.sleep(throttle_sec)
        return followees


# ----------------------------- Exporters -----------------------------

def to_dataframe(items: list[Followee]) -> pd.DataFrame:
    rows = []
    for x in items:
        verified_status = "YES" if x.is_verified else "NO" if x.is_verified is not None else None
        private_status = "YES" if x.is_private else "NO" if x.is_private is not None else None
        
        rows.append({
            "Username": x.username,
            "Profile Link": x.profile_link,
            "Full Name": x.full_name,
            "Verified": verified_status,
            "Private": private_status,
            "User ID": x.user_id,
            "Profile Pic URL": x.profile_pic_url,
            "Followers": x.followers,
            "Followings": x.followings,
            "Bio": x.bio,
        })
    df = pd.DataFrame(rows)
    return df


def export_csv_json_excel(df: pd.DataFrame, out_dir: str, base_name: str,
                          embed_thumbs: bool = False, thumb_size: int = 64):
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = safe_filename(f"{base_name}_{ts}")

    csv_path = os.path.join(out_dir, f"{stem}.csv")
    json_path = os.path.join(out_dir, f"{stem}.json")
    xlsx_path = os.path.join(out_dir, f"{stem}.xlsx")

    df.to_csv(csv_path, index=False, encoding='utf-8')
    df.to_json(json_path, orient='records', force_ascii=False, indent=2)

    if Workbook is None:
        # Fallback: use pandas Excel without thumbnails/hyperlinks formatting
        try:
            df.to_excel(xlsx_path, index=False)
        except Exception:
            pass
        return csv_path, json_path, xlsx_path

    # Pretty Excel with hyperlinks, optional thumbnails
    wb = Workbook()
    ws = wb.active
    ws.title = "Followings"

    # Headers
    headers = list(df.columns)
    ws.append(headers)

    # Column widths baseline
    col_widths = {i+1: max(10, len(h)+2) for i, h in enumerate(headers)}

    # Pre-download thumbnails if needed
    images_cache = {}
    if embed_thumbs and 'Profile Pic URL' in df.columns:
        sess = requests.Session()
        for idx, url in enumerate(df['Profile Pic URL']):
            if not isinstance(url, str) or not url:
                images_cache[idx] = None
                continue
            try:
                r = sess.get(url, timeout=20)
                if r.status_code == 200:
                    images_cache[idx] = r.content
                else:
                    images_cache[idx] = None
            except Exception:
                images_cache[idx] = None

    # Rows
    for i, row in df.iterrows():
        excel_row = []
        for h in headers:
            val = row.get(h)
            if h == 'Profile Link' and isinstance(val, str):
                excel_row.append(val)
            else:
                excel_row.append(val)
        ws.append(excel_row)
        # Set hyperlink in cell (Profile Link)
        try:
            link_col = headers.index('Profile Link') + 1
            c = ws.cell(row=i+2, column=link_col)
            c.hyperlink = c.value
            c.style = 'Hyperlink'
        except Exception:
            pass

        # Embed thumbnail image in a dedicated column if requested
        if embed_thumbs and 'Profile Pic URL' in headers:
            try:
                img_bytes = images_cache.get(i)
                if img_bytes:
                    # Save to bytes -> PIL not required; openpyxl needs a file-like
                    tmp_path = None
                    try:
                        tmp_path = os.path.join(out_dir, f"._thumb_{i}.png")
                        with open(tmp_path, 'wb') as fh:
                            fh.write(img_bytes)
                        xlimg = XLImage(tmp_path)
                        # Resize roughly by setting height via row dimension
                        ws.row_dimensions[i+2].height = thumb_size
                        pic_col = headers.index('Profile Pic URL') + 1
                        # Insert image to the left of URL (same cell)
                        ws.add_image(xlimg, f"{get_column_letter(pic_col)}{i+2}")
                    finally:
                        if tmp_path and os.path.exists(tmp_path):
                            try:
                                os.remove(tmp_path)
                            except Exception:
                                pass
            except Exception:
                pass

        # Track max width
        for col_idx, h in enumerate(headers, start=1):
            v = row.get(h)
            vlen = len(str(v)) if v is not None else 0
            col_widths[col_idx] = min(60, max(col_widths[col_idx], vlen + 2))

    # Apply column widths
    for col_idx, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    wb.save(xlsx_path)
    return csv_path, json_path, xlsx_path


def load_last_snapshot(path: str) -> set[str]:
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return set(data)
    except Exception:
        return set()


def save_snapshot(path: str, usernames: list[str]):
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(sorted(usernames), f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ----------------------------- GUI -----------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Instagram Followings Exporter (cookies.txt)")
        self.geometry("780x540")
        self.minsize(720, 520)

        self.cookies_path = tk.StringVar()
        self.output_dir = tk.StringVar(value=os.getcwd())

        self.opt_verified_only = tk.BooleanVar(value=False)
        self.opt_private_only = tk.BooleanVar(value=False)
        self.opt_fetch_details = tk.BooleanVar(value=True)
        self.opt_embed_thumbs = tk.BooleanVar(value=False)

        self.status = tk.StringVar(value="Ready.")

        self._build()
        self._enable_drag_drop()

    def _build(self):
        pad = 10
        main = ttk.Frame(self)
        main.pack(fill=tk.BOTH, expand=True, padx=pad, pady=pad)

        # Row 1: cookies + browse
        frm1 = ttk.LabelFrame(main, text="Authentication")
        frm1.pack(fill=tk.X, pady=(0, pad))

        ttk.Label(frm1, text="cookies.txt path:").grid(row=0, column=0, sticky='w', padx=8, pady=6)
        self.cookies_entry = ttk.Entry(frm1, textvariable=self.cookies_path, width=64)
        self.cookies_entry.grid(row=0, column=1, sticky='we', padx=8, pady=6)
        ttk.Button(frm1, text="Browse…", command=self.browse_cookies).grid(row=0, column=2, padx=8, pady=6)

        ttk.Label(frm1, text="💡 Tip: You can drag & drop cookies.txt file here", 
                 font=('TkDefaultFont', 8), foreground='gray').grid(row=1, column=1, sticky='w', padx=8, pady=(0, 6))

        frm1.columnconfigure(1, weight=1)

        # Output dir
        frm2 = ttk.LabelFrame(main, text="Output")
        frm2.pack(fill=tk.X, pady=(0, pad))
        ttk.Label(frm2, text="Output folder:").grid(row=0, column=0, sticky='w', padx=8, pady=6)
        ttk.Entry(frm2, textvariable=self.output_dir, width=64).grid(row=0, column=1, sticky='we', padx=8, pady=6)
        ttk.Button(frm2, text="Choose…", command=self.browse_output).grid(row=0, column=2, padx=8, pady=6)

        frm2.columnconfigure(1, weight=1)

        # Options
        frm3 = ttk.LabelFrame(main, text="Options")
        frm3.pack(fill=tk.X, pady=(0, pad))
        ttk.Checkbutton(frm3, text="Verified only", variable=self.opt_verified_only).grid(row=0, column=0, sticky='w', padx=8, pady=6)
        ttk.Checkbutton(frm3, text="Private only", variable=self.opt_private_only).grid(row=0, column=1, sticky='w', padx=8, pady=6)
        ttk.Checkbutton(frm3, text="Fetch stats & bio (slower)", variable=self.opt_fetch_details).grid(row=0, column=2, sticky='w', padx=8, pady=6)
        ttk.Checkbutton(frm3, text="Embed profile thumbnails in Excel", variable=self.opt_embed_thumbs).grid(row=0, column=3, sticky='w', padx=8, pady=6)

        # Actions
        frm4 = ttk.Frame(main)
        frm4.pack(fill=tk.X, pady=(0, pad))
        ttk.Button(frm4, text="Download Followings", command=self.run_export).pack(side=tk.LEFT, padx=4)
        ttk.Button(frm4, text="Validate Session", command=self.validate_session).pack(side=tk.LEFT, padx=4)

        # Status + Log
        self.txt = tk.Text(main, height=16)
        self.txt.pack(fill=tk.BOTH, expand=True)

        self.statusbar = ttk.Label(self, textvariable=self.status, anchor='w')
        self.statusbar.pack(fill=tk.X, side=tk.BOTTOM)

    def _enable_drag_drop(self):
        def drop_enter(event):
            event.widget.focus_force()
            return 'copy'

        def drop_position(event):
            return 'copy'

        def drop_leave(event):
            return 'copy'

        def drop(event):
            try:
                # Handle different data formats
                if hasattr(event, 'data'):
                    files = event.data
                else:
                    files = event.widget.tk.splitlist(event.widget.tk.call('tkdnd::GetDroppedData'))
                
                if files:
                    file_path = files[0] if isinstance(files, (list, tuple)) else str(files)
                    # Remove curly braces if present (Windows format)
                    file_path = file_path.strip('{}')
                    
                    if file_path.lower().endswith('.txt'):
                        self.cookies_path.set(file_path)
                        self.log(f"[+] Cookies file loaded via drag & drop: {os.path.basename(file_path)}")
                    else:
                        self.log("[!] Please drop a .txt file (cookies.txt)")
            except Exception as e:
                self.log(f"[!] Error handling dropped file: {e}")
            return 'copy'

        # Enable drag and drop on the cookies entry widget
        try:
            # Try to use tkinterdnd2 if available
            try:
                import tkinterdnd2 as tkdnd
                self.cookies_entry.drop_target_register(tkdnd.DND_FILES)
                self.cookies_entry.dnd_bind('<<Drop>>', drop)
                self.cookies_entry.dnd_bind('<<DropEnter>>', drop_enter)
                self.cookies_entry.dnd_bind('<<DropPosition>>', drop_position)
                self.cookies_entry.dnd_bind('<<DropLeave>>', drop_leave)
            except ImportError:
                # Fallback: bind to standard tkinter events for basic file handling
                def handle_file_drop(event):
                    # This won't work for actual drag-drop but provides a placeholder
                    pass
                
                # Bind to button-1 for click to browse as fallback
                self.cookies_entry.bind('<Button-3>', lambda e: self.browse_cookies())
                self.log("[!] Advanced drag & drop not available. Right-click entry field to browse for file.")
                
        except Exception:
            # If all else fails, just provide the browse button functionality
            pass

    def browse_cookies(self):
        path = filedialog.askopenfilename(title="Select cookies.txt", filetypes=[("Text", "*.txt"), ("All", "*.*")])
        if path:
            self.cookies_path.set(path)

    def browse_output(self):
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            self.output_dir.set(path)

    def log(self, msg: str):
        self.txt.insert(tk.END, msg + "\n")
        self.txt.see(tk.END)
        self.status.set(msg)
        self.update_idletasks()

    def validate_inputs(self):
        if not self.cookies_path.get() or not os.path.exists(self.cookies_path.get()):
            messagebox.showerror("Missing", "cookies.txt ka sahi path select karein ya file ko drag & drop karein.")
            return False
        if not os.path.isdir(self.output_dir.get()):
            messagebox.showerror("Missing", "Valid output folder select karein.")
            return False
        return True

    def validate_session(self):
        if not self.validate_inputs():
            return
        try:
            ig = IGClient(self.cookies_path.get())
            ok, username = ig.validate_session()
            if ok:
                if username:
                    self.log(f"[+] Session valid hai. Logged in as: @{username}")
                else:
                    self.log("[+] Session valid hai (current_user OK).")
            else:
                self.log("[!] Session invalid ya expired hai. Browser se cookies dobara export karein.")
        except Exception as e:
            self.log(f"[!] Validation error: {e}")

    def run_export(self):
        if not self.validate_inputs():
            return
        threading.Thread(target=self._run_export_worker, daemon=True).start()

    def _run_export_worker(self):
        try:
            self.log("[•] Session initialize ho rahi hai…")
            ig = IGClient(self.cookies_path.get())

            self.log("[•] Session validate kar rahe hain aur username fetch kar rahe hain…")
            session_valid, username = ig.validate_session()
            if not session_valid:
                self.log("[!] Session invalid ya expired hai. Browser se cookies dobara export karein.")
                return
            if not username:
                self.log("[!] Username fetch nahi ho paya session se.")
                return
            
            self.log(f"[+] Logged in as: @{username}")

            self.log("[•] User ID nikal rahe hain…")
            user_id = ig.get_user_id(username)
            self.log(f"[+] User ID: {user_id}")

            fetch_details = self.opt_fetch_details.get()
            self.log("[•] Followings fetch ho rahe hain (paging handled)…")
            followees = ig.get_followings(user_id, fetch_details=fetch_details)
            self.log(f"[+] Total followings: {len(followees)}")

            # Filters
            if self.opt_verified_only.get():
                followees = [f for f in followees if f.is_verified]
                self.log(f"[•] Verified-only filter applied → {len(followees)}")
            if self.opt_private_only.get():
                followees = [f for f in followees if f.is_private]
                self.log(f"[•] Private-only filter applied → {len(followees)}")

            # DataFrame
            df = to_dataframe(followees)

            # Change tracking
            snapshot_path = os.path.join(self.output_dir.get(), f"_last_followings_{safe_filename(username)}.json")
            prev = load_last_snapshot(snapshot_path)
            current = set(df['Username'].dropna().tolist())
            added = sorted(list(current - prev))
            removed = sorted(list(prev - current))

            if prev:
                self.log(f"[•] Change tracking: +{len(added)} added, -{len(removed)} removed")

            # Export
            self.log("[•] Exporting CSV/JSON/Excel…")
            csv_path, json_path, xlsx_path = export_csv_json_excel(
                df, self.output_dir.get(), base_name=f"followings_{username}",
                embed_thumbs=self.opt_embed_thumbs.get()
            )

            # Save snapshot
            save_snapshot(snapshot_path, sorted(current))

            # Write change files if any
            ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            if added:
                pd.DataFrame(added, columns=['Username']).to_csv(
                    os.path.join(self.output_dir.get(), f"added_{ts}.csv"), index=False, encoding='utf-8')
            if removed:
                pd.DataFrame(removed, columns=['Username']).to_csv(
                    os.path.join(self.output_dir.get(), f"removed_{ts}.csv"), index=False, encoding='utf-8')

            self.log("[✓] Done. Files:")
            self.log(f"    CSV : {csv_path}")
            self.log(f"    JSON: {json_path}")
            self.log(f"    XLSX: {xlsx_path}")
            if added or removed:
                self.log("    (Change tracking CSVs bhi output folder me save ho gayi hain.)")

        except Exception as e:
            self.log(f"[!] Error: {e}")


if __name__ == '__main__':
    app = App()
    app.mainloop()
